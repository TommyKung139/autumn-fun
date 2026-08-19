# Reconciles data/reference.json's combined_layout with the 分桌表 sheet.
#
# Why this is needed: the floor-plan sheet (分桌座位圖請參考) and the seating
# assignment sheet (分桌表) label the same tables slightly differently -
# the floor plan writes 'A12 A13' where the assignment sheet writes
# 'A12A & A13', and 'C8 C9' where the assignment writes 'C8、C9'. Employees are
# told the 分桌表 code on their itinerary card, so the map has to use that same
# code or their table cannot be found/highlighted on the map.
#
# This script rewrites each map item's `code`, `count` and `dept_raw` from 分桌表
# (the authoritative, most up-to-date source for who actually sits where), while
# leaving the geometry (col/row/spans, drawn from the venue's official floor
# plan) untouched.
#
# Usage: SRC_XLSX=/path/to/workbook.xlsx python3 scripts/sync_layout_codes.py
import openpyxl, json, os, re

SRC = os.environ.get('SRC_XLSX')
HERE = os.path.dirname(os.path.abspath(__file__))
REF = os.path.join(HERE, '..', 'data', 'reference.json')

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['分桌表']

# 分桌表 -> {code: (人數, 科別)}
assign = {}
for r in range(4, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    if not code or re.match(r'^[ABC]區', str(code)) or '板前' in str(code):
        continue
    assign[str(code).strip()] = (ws.cell(row=r, column=2).value,
                                 str(ws.cell(row=r, column=3).value or '').strip())
print(f'分桌表: {len(assign)} table codes')

# Current map code -> the 分桌表 code naming the same physical group.
RENAME = {
    'A11 A12': 'A11 & A12',
    'A12 A13': 'A12A & A13',
    'A16 A17': 'A16 & A17',
    'A17 A18': 'A17A & A18',
    'A27 A28': 'A27 & A28',
    'A28 A29': 'A28A & A29',
    'A30 A31': 'A30 & A31',
    'A31 A32': 'A31A & A32',
    'A38': 'A38&A38A',
    'C6 C7': 'C6、C7',
    'C8 C9': 'C8、C9',
    'C10 C11': 'C10、C11',
    'C13 C14': 'C13、C14',
    'C15 C16': 'C15、C16',
    'C17 C18': 'C17、C18',
}

ref = json.load(open(REF, encoding='utf-8'))
items = ref['combined_layout']['items']

new_items = []
unknown = []
for it in items:
    if it.get('type') != 'table':
        new_items.append(it)
        continue
    code = it['code']
    if code == '板前':
        new_items.append(it)
        continue
    # 'A7 A8' is one box on the floor plan but two separate tables in 分桌表 -
    # split it so each table carries its own real occupancy.
    if code == 'A7 A8':
        for i, sub in enumerate(['A7', 'A8']):
            cnt, dept = assign.get(sub, (it.get('count'), it.get('dept_raw')))
            new_items.append(dict(it, code=sub, count=cnt, dept_raw=dept,
                                  row=it['row'] + i * 2, row_span=2))
        continue
    target = RENAME.get(code, code)
    if target not in assign:
        unknown.append(code)
        new_items.append(it)
        continue
    cnt, dept = assign[target]
    new_items.append(dict(it, code=target, count=cnt, dept_raw=dept))

ref['combined_layout']['items'] = new_items
json.dump(ref, open(REF, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

mapped = {i['code'] for i in new_items if i.get('type') == 'table' and i['code'] != '板前'}
print(f'map table items: {len(mapped)}')
print('map codes NOT in 分桌表:', sorted(mapped - set(assign)) or 'none')
print('分桌表 codes NOT on map:', sorted(set(assign) - mapped) or 'none')
if unknown:
    print('left unchanged (no 分桌表 match):', unknown)
