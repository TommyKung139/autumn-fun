# Independent re-derivation of every employee field from the source workbook,
# used to audit data/employees.json for extraction errors. Read-only: prints a
# report, never writes. Run: SRC_XLSX=<path> python3 scripts/audit_extraction.py
import openpyxl, re, json, os, collections, unicodedata

SRC = os.environ.get('SRC_XLSX')
wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------- roster ----------------
ws = wb['(彙整)報名表(員工)']
roster = {}       # emp_id -> record
for row in ws.iter_rows(min_row=3, values_only=True):
    emp_id, name, attend = row[5], row[6], row[8]
    if not emp_id:
        continue
    if not attend or not str(attend).startswith('參加'):
        continue
    roster[emp_id] = {
        'emp_id': emp_id, 'name': name, 'division': row[2], 'department': row[3],
        'region': row[4], 'emp_type': row[7], 'batch': row[0], 'meal': row[13],
        'outbound_station': row[16], 'return_station': row[17],
        'afternoon_activity': row[18], 'evening_activity': row[19],
    }
print(f'ROSTER: {len(roster)} attending employees')

# Division -> short tag used in the assignment sheets
DIV_TAG = {'經營策略處': '經策', '數位平台經營處': '數平', '個金作業暨資訊處': '作資'}
for e in roster.values():
    e['tag'] = DIV_TAG.get(e['division'])

# ---------------- name matching ----------------
def nfkc(s):
    return unicodedata.normalize('NFKC', str(s)) if s is not None else ''

# Characters that are the same person but typed differently in different sheets.
CHAR_FIX = str.maketrans({'豔': '艷', '鳳': '鳯'})

def canon(s):
    """Aggressively normalise a name for comparison: full/half-width unified,
    all whitespace removed, known variant characters folded together."""
    s = nfkc(s).translate(CHAR_FIX)
    return re.sub(r'\s+', '', s)

DEP_MARKERS = ('眷屬', '眷)', '眷）', '(眷', '（眷')

def is_dependent_cell(raw):
    r = nfkc(raw)
    return any(m in r for m in ('眷屬', '(眷)', '（眷）', '(眷', '（眷'))

def base_and_tag(raw):
    """Split 'X(Y)' -> ('X','Y'); leaves 'X' alone. Only strips ONE trailing group."""
    r = nfkc(raw).strip()
    m = re.match(r'^(.*?)[（(]([^)）]*)[)）]\s*$', r, re.S)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return r, None

# index by canonical FULL name (i.e. including any '(Chih)' the roster itself carries)
full_index = collections.defaultdict(list)
# index by canonical BASE name (parenthetical stripped) - may be ambiguous
base_index = collections.defaultdict(list)
for eid, e in roster.items():
    full_index[canon(e['name'])].append(eid)
    b, _ = base_and_tag(e['name'])
    base_index[canon(b)].append(eid)

def match(raw, dept_hint=None):
    """Resolve an assignment-sheet cell to an emp_id.
    Returns (emp_id, status) where status is one of:
      'exact' | 'base' | 'disambiguated' | 'dependent' | 'ambiguous' | 'unknown'
    """
    if raw is None or not str(raw).strip():
        return None, 'empty'
    if is_dependent_cell(raw):
        return None, 'dependent'
    c = canon(raw)
    # 1. exact full-name hit (this is what catches '林芝萱(Chih)')
    if c in full_index and len(full_index[c]) == 1:
        return full_index[c][0], 'exact'
    if c in full_index and len(full_index[c]) > 1:
        cands = full_index[c]
    else:
        b, _ = base_and_tag(raw)
        cands = base_index.get(canon(b), [])
    if not cands:
        return None, 'unknown'
    if len(cands) == 1:
        return cands[0], 'base'
    # ambiguous - try the department hint
    if dept_hint:
        narrowed = [eid for eid in cands if roster[eid]['tag'] == dept_hint]
        if len(narrowed) == 1:
            return narrowed[0], 'disambiguated'
    return None, 'ambiguous'

problems = collections.defaultdict(list)

def note(kind, msg):
    problems[kind].append(msg)

# ---------------- bus tables ----------------
ws3 = wb['分車表 (2)']

def dept_from_header(text):
    for tag in ('經策', '數平', '作資'):
        if tag in str(text):
            return tag
    return None

def parse_bus(header_row, unit_row, row_start, row_end, field):
    """unit_row: extra row holding department counts (table 2 only); None for table 1."""
    cols = {}
    for c in range(2, ws3.max_column + 1):
        v = ws3.cell(row=header_row, column=c).value
        if v and isinstance(v, str) and '車' in v:
            label = v.split('\n')[0].strip()
            hint_src = v if unit_row is None else ws3.cell(row=unit_row, column=c).value
            # a header naming 2+ departments cannot disambiguate anything
            tags = [t for t in ('經策', '數平', '作資') if t in str(hint_src)]
            cols[c] = (label, tags[0] if len(tags) == 1 else None)
    out = {}
    for r in range(row_start, row_end + 1):
        for c, (label, hint) in cols.items():
            v = ws3.cell(row=r, column=c).value
            eid, status = match(v, hint)
            if eid:
                if eid in out and out[eid] != label:
                    note('bus_conflict', f'{field}: {roster[eid]["name"]} ({eid}) appears in both {out[eid]} and {label}')
                out[eid] = label
            elif status in ('ambiguous', 'unknown'):
                note(f'{field}_unmatched', f'r{r}c{c} ({label}) {v!r} -> {status}')
    return out

morning = parse_bus(4, None, 5, 46, 'morning_bus')
activity = parse_bus(48, 49, 50, 90, 'activity_bus')
print(f'BUS: morning matched {len(morning)}, activity matched {len(activity)}')

# ---------------- table assignment ----------------
ws4 = wb['分桌表']
tables = {}
for r in range(4, ws4.max_row + 1):
    table_no = ws4.cell(row=r, column=1).value
    if not table_no or re.match(r'^[ABC]區', str(table_no)) or '板前' in str(table_no):
        continue
    hint = dept_from_header(ws4.cell(row=r, column=3).value)
    for c in range(4, 12):
        v = ws4.cell(row=r, column=c).value
        eid, status = match(v, hint)
        if eid:
            if eid in tables and tables[eid] != table_no:
                note('table_conflict', f'{roster[eid]["name"]} ({eid}) in both {tables[eid]} and {table_no}')
            tables[eid] = str(table_no).strip()
        elif status in ('ambiguous', 'unknown'):
            note('table_unmatched', f'r{r}c{c} ({table_no}) {v!r} -> {status}')
print(f'TABLE: matched {len(tables)}')

# ---------------- HSR ----------------
hsr_raw = json.load(open(os.path.join(os.path.dirname(__file__), 'hsr_raw.json')))
def attach(records, field):
    out = {}
    for rec in records:
        hint = dept_from_header(rec.get('dept_raw'))
        eid, status = match(rec['name_raw'], hint)
        if eid:
            if eid in out:
                note(f'{field}_conflict', f'{roster[eid]["name"]} ({eid}) has two {field} seats')
            out[eid] = {
                'train_no': rec['train_no'], 'car': rec['car'],
                'seat_letter': rec['seat_letter'], 'row_idx': rec['row_idx'],
                'station': rec['station'], 'note': rec['note'],
            }
        elif status in ('ambiguous', 'unknown'):
            note(f'{field}_unmatched', f'{rec["name_raw"]!r} {rec.get("dept_raw")!r} -> {status}')
    return out
hsr_go = attach(hsr_raw['go'], 'hsr_outbound')
hsr_back = attach(hsr_raw['back'], 'hsr_return')
print(f'HSR: outbound {len(hsr_go)}, return {len(hsr_back)}')

# ---------------- dependents ----------------
ws2 = wb['(彙整)報名表(眷屬)']
def normalize_emp_id(raw_id):
    if not raw_id or raw_id in roster:
        return raw_id
    m = re.match(r'^([A-Za-z]+)(\d+)$', str(raw_id).strip())
    if not m:
        return raw_id
    padded = m.group(1).upper() + m.group(2).zfill(8)
    return padded if padded in roster else raw_id

deps = collections.defaultdict(list)
dep_unmatched = []
for row in ws2.iter_rows(min_row=4, values_only=True):
    emp_name, emp_id = row[4], row[5]
    if not emp_id:
        continue
    emp_id = normalize_emp_id(emp_id)
    if emp_id not in roster:
        dep_unmatched.append((emp_id, emp_name, row[6]))
        continue
    deps[emp_id].append({
        'name': row[6], 'relation': row[7], 'age_band': row[17],
        'outbound_station': row[18], 'return_station': row[19],
        'afternoon_activity': row[20], 'evening_activity': row[21],
    })
print(f'DEPENDENTS: matched {sum(len(v) for v in deps.values())} rows onto {len(deps)} employees, unmatched {len(dep_unmatched)}')

# ---------------- build expected records & diff ----------------
expected = {}
for eid, e in roster.items():
    ws_row = None
    expected[eid] = {
        'emp_id': eid, 'name': e['name'], 'division': e['division'],
        'department': e['department'], 'region': e['region'], 'emp_type': e['emp_type'],
        'batch': e['batch'], 'meal': e['meal'],
        'outbound_station': e['outbound_station'], 'return_station': e['return_station'],
        'afternoon_activity': e['afternoon_activity'], 'evening_activity': e['evening_activity'],
        'morning_bus': morning.get(eid), 'activity_bus': activity.get(eid),
        'table': tables.get(eid), 'hsr_outbound': hsr_go.get(eid), 'hsr_return': hsr_back.get(eid),
        'dependents': deps.get(eid, []),
    }

current = {r['emp_id']: r for r in json.load(open(os.path.join(os.path.dirname(__file__), '..', 'data', 'employees.json')))}

print('\n' + '=' * 70)
print('DIFF vs current data/employees.json')
print('=' * 70)
only_cur = set(current) - set(expected)
only_exp = set(expected) - set(current)
if only_cur: print('IN CURRENT BUT NOT EXPECTED:', only_cur)
if only_exp: print('IN EXPECTED BUT NOT CURRENT:', only_exp)

FIELDS = ['morning_bus', 'activity_bus', 'table', 'hsr_outbound', 'hsr_return',
          'meal', 'outbound_station', 'return_station', 'afternoon_activity',
          'evening_activity', 'division', 'department', 'region', 'emp_type']
diff_count = collections.Counter()
for eid in sorted(set(expected) & set(current)):
    exp, cur = expected[eid], current[eid]
    for f in FIELDS:
        if exp[f] != cur.get(f):
            diff_count[f] += 1
            print(f'  {eid} {exp["name"]:<8} {f:<16} current={cur.get(f)!r:<28} expected={exp[f]!r}')
    if len(exp['dependents']) != len(cur.get('dependents', [])):
        diff_count['dependents'] += 1
        print(f'  {eid} {exp["name"]:<8} dependents      current={len(cur.get("dependents", []))} expected={len(exp["dependents"])}')

print('\nDIFF SUMMARY:', dict(diff_count) or 'no differences')

print('\n' + '=' * 70)
print('EXTRACTION PROBLEMS (names in assignment sheets that resolve to nobody)')
print('=' * 70)
for kind, msgs in sorted(problems.items()):
    print(f'\n--- {kind} ({len(msgs)}) ---')
    for m in msgs:
        print('   ', m)

print('\n--- unmatched dependent rows (%d) ---' % len(dep_unmatched))
for u in dep_unmatched:
    print('   ', u)

# coverage report
print('\n' + '=' * 70)
print('COVERAGE (expected)')
print('=' * 70)
for f in ['morning_bus', 'activity_bus', 'table', 'hsr_outbound', 'hsr_return']:
    n = sum(1 for e in expected.values() if e[f])
    print(f'  {f:<16} {n}/{len(expected)}')
