# One-off script that generates ../data/table_layout.json (the "島語 午宴座位圖" grid
# data shown on info.html) from the "分桌座位圖請參考" sheet in the original registration
# workbook. Not run at deploy time - kept for provenance / so it can be regenerated if the
# seating plan changes.
# Requires: pip install openpyxl
import openpyxl, re, json, os

SRC = os.environ.get('SRC_XLSX', './2026秋郊報名表_第二梯_數平經策作資0818.xlsx')
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['分桌座位圖請參考']

merge_map = {}
for mc in ws.merged_cells.ranges:
    for r in range(mc.min_row, mc.max_row + 1):
        for c in range(mc.min_col, mc.max_col + 1):
            merge_map[(r, c)] = (mc.min_row, mc.min_col, mc.max_row, mc.max_col)

seen = set()
tables = []
table_pat = re.compile(r'^([A-C]\d+(?:[\s、\.,，]+[A-C]?\d+)*|板前)\s*\n\((\d+)\)\s*\n(.+)$')

for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if not v or not isinstance(v, str):
            continue
        key = (cell.row, cell.column)
        rng = merge_map.get(key, (cell.row, cell.column, cell.row, cell.column))
        top = (rng[0], rng[1])
        if top in seen:
            continue
        m = table_pat.match(v.strip())
        if not m:
            continue
        seen.add(top)
        code, count, dept = m.group(1), m.group(2), m.group(3)
        tables.append({
            'code': code.replace('\n', ' ').strip(),
            'count': int(count),
            'dept_raw': dept.strip(),
            'row': rng[0], 'col': rng[1],
            'row_span': rng[2] - rng[0] + 1, 'col_span': rng[3] - rng[1] + 1,
        })

rows = [t['row'] for t in tables] + [t['row'] + t['row_span'] - 1 for t in tables]
cols = [t['col'] for t in tables] + [t['col'] + t['col_span'] - 1 for t in tables]
min_row, min_col = min(rows), min(cols)
for t in tables:
    t['grid_row'] = t['row'] - min_row + 1
    t['grid_col'] = t['col'] - min_col + 1

out_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'reference.json')
ref = json.load(open(out_path, encoding='utf-8'))
ref['table_layout'] = {
    'grid_rows': max(rows) - min_row + 1,
    'grid_cols': max(cols) - min_col + 1,
    'tables': tables,
}
json.dump(ref, open(out_path, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
print('table_layout written into data/reference.json, tables:', len(tables))
