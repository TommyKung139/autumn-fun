# Data-extraction script that generates ../data/employees.json from the
# "2026秋郊報名表_第二梯_數平經策作資0818.xlsx" registration workbook.
# Not run at deploy time / not needed for the website to work - kept so the data
# can be regenerated if the Excel is updated.
#
# Requires: pip install openpyxl
# Also requires scripts/hsr_raw.json (extracted separately from the two HSR
# seating sheets) to be present in this folder before running.
#
# Usage:  SRC_XLSX=/path/to/workbook.xlsx python3 scripts/build_employees.py
#
# ---------------------------------------------------------------------------
# Name matching notes (this is where the hard parts are)
#
# The registration sheet is the source of truth for WHO is attending. The bus /
# table / HSR sheets refer to those people by NAME ONLY, so every assignment has
# to be matched back by name. Four things make that non-trivial, and each one
# caused a real mis-assignment in an earlier version of this script:
#
#  1. Two attending employees share the name 陳怡君 (經策/數據營運部 and
#     數平/數位平台部). Blindly taking the first match put one person's whole
#     itinerary on the other. Resolved with department hints from the sheet
#     headers, then by cross-checking the afternoon/evening activity choice.
#  2. Two employees are called 林芝萱 and the ROSTER itself disambiguates them
#     inline: '林芝萱(Chih)' and '林芝萱 (Boa)'. Stripping the parenthetical
#     before lookup made both of them match nothing at all, so both had a
#     completely empty itinerary. Now an exact full-name match is tried first.
#  3. Dependents appear in the seating sheets as '王小明(眷)' / '王小明（眷屬）'
#     / '王小明眷屬'. The '(眷)' short form was not being filtered, so a
#     dependent's HSR seat overwrote the employee's own seat. All dependent
#     markers are now filtered out (dependents come from the 眷屬 sheet).
#  4. The same person is spelled with variant characters across sheets
#     (莊若艷 / 莊若豔). Names are compared after NFKC normalisation, whitespace
#     removal and folding of the known variant pairs.
# ---------------------------------------------------------------------------
import openpyxl, re, json, os, collections, unicodedata

SRC = os.environ.get('SRC_XLSX', './2026秋郊報名表_第二梯_數平經策作資0818.xlsx')
HERE = os.path.dirname(os.path.abspath(__file__))
wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------- 1. Roster (source of truth for who is attending) ----------
ws = wb['(彙整)報名表(員工)']
employees = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    batch, seq, division, dept, region, emp_id, name, emp_type, attend = row[0:9]
    if not emp_id or not attend or not str(attend).startswith('參加'):
        continue
    fam_adult, fam_6_12, fam_3_6, fam_under3 = row[21], row[22], row[23], row[24]
    employees[emp_id] = {
        'emp_id': emp_id, 'name': name, 'division': division, 'department': dept,
        'region': region, 'emp_type': emp_type, 'batch': batch, 'meal': row[13],
        'outbound_station': row[16], 'return_station': row[17],
        'afternoon_activity': row[18], 'evening_activity': row[19],
        'bring_family': (row[20] == '是'),
        'family_summary': {'adult': fam_adult or 0, 'c6_12': fam_6_12 or 0,
                           'c3_6': fam_3_6 or 0, 'under3': fam_under3 or 0},
        'morning_bus': None, 'activity_bus': None, 'table': None,
        'hsr_outbound': None, 'hsr_return': None, 'dependents': []
    }

DIV_TAG = {'經營策略處': '經策', '數位平台經營處': '數平', '個金作業暨資訊處': '作資'}
for e in employees.values():
    e['_tag'] = DIV_TAG.get(e['division'])

# ---------- 2. Name matching ----------
# Same person, different character across sheets.
CHAR_FIX = str.maketrans({'豔': '艷', '鳳': '鳯'})

def canon(s):
    """Normalise a name for comparison: NFKC (unifies full/half-width brackets and
    spaces), drop all whitespace, fold known variant characters."""
    if s is None:
        return ''
    return re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(s)).translate(CHAR_FIX))

def is_dependent_cell(raw):
    """True for cells naming a dependent rather than the employee, in any of the
    forms used across the sheets: (眷) / （眷） / (眷屬) / 眷屬 / (眷屬1)."""
    r = unicodedata.normalize('NFKC', str(raw))
    return '眷' in r

def base_and_tag(raw):
    """'林芝萱(Chih)' -> ('林芝萱', 'Chih'); '王小明' -> ('王小明', None)."""
    r = unicodedata.normalize('NFKC', str(raw)).strip()
    m = re.match(r'^(.*?)\(([^)]*)\)\s*$', r, re.S)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return r, None

full_index = collections.defaultdict(list)   # canonical full roster name -> [emp_id]
base_index = collections.defaultdict(list)   # canonical name minus parenthetical
for eid, e in employees.items():
    full_index[canon(e['name'])].append(eid)
    b, _ = base_and_tag(e['name'])
    base_index[canon(b)].append(eid)

unresolved = []

def match_emp_id(raw, dept_hint=None, activity_hint=None, where=''):
    """Resolve an assignment-sheet cell to an emp_id, or None.

    dept_hint:     '經策' / '數平' / '作資' taken from the sheet's column header.
    activity_hint: (afternoon, evening) strings implied by the column, used to
                   break a tie the department hint cannot (e.g. a bus carrying
                   two departments)."""
    if raw is None or not str(raw).strip():
        return None
    if is_dependent_cell(raw):
        return None
    c = canon(raw)
    # Exact full-name match first: this is what correctly resolves the roster's
    # own inline disambiguation, e.g. '林芝萱(Chih)' vs '林芝萱 (Boa)'.
    cands = full_index.get(c)
    if not cands:
        b, _ = base_and_tag(raw)
        cands = base_index.get(canon(b), [])
    if not cands:
        unresolved.append((where, str(raw), 'no roster match'))
        return None
    if len(cands) == 1:
        return cands[0]
    if dept_hint:
        narrowed = [eid for eid in cands if employees[eid]['_tag'] == dept_hint]
        if len(narrowed) == 1:
            return narrowed[0]
        if narrowed:
            cands = narrowed
    if activity_hint:
        aft, eve = activity_hint
        narrowed = [eid for eid in cands
                    if (aft is None or employees[eid]['afternoon_activity'] == aft)
                    and (eve is None or employees[eid]['evening_activity'] == eve)]
        if len(narrowed) == 1:
            return narrowed[0]
    unresolved.append((where, str(raw), f'ambiguous between {cands}'))
    return None

# ---------- 3. Dependents ----------
ws2 = wb['(彙整)報名表(眷屬)']

def normalize_emp_id(raw_id):
    """Fix employee-ID typos that drop a leading zero ('Z0034011' -> 'Z00034011').
    Only applied when the zero-padded ID actually exists, so it never invents a
    match."""
    if not raw_id or raw_id in employees:
        return raw_id
    m = re.match(r'^([A-Za-z]+)(\d+)$', str(raw_id).strip())
    if not m:
        return raw_id
    padded = m.group(1).upper() + m.group(2).zfill(8)
    return padded if padded in employees else raw_id

dep_matched, dep_unmatched = 0, []
for row in ws2.iter_rows(min_row=4, values_only=True):
    emp_name, emp_id = row[4], row[5]
    if not emp_id:
        continue
    emp_id = normalize_emp_id(emp_id)
    if emp_id not in employees:
        dep_unmatched.append((emp_id, emp_name, row[6]))
        continue
    dep_matched += 1
    employees[emp_id]['dependents'].append({
        'name': row[6], 'relation': row[7], 'age_band': row[17],
        'outbound_station': row[18], 'return_station': row[19],
        'afternoon_activity': row[20], 'evening_activity': row[21],
    })
print(f'dependents: matched {dep_matched} rows, unmatched {len(dep_unmatched)}')

# ---------- 4. Bus tables ----------
# 分車表一 (header row 4, rows 5-46)  = morning outbound coach
# 分車表二 (header row 48 + unit row 49, rows 50-90) = afternoon activity coach
ws3 = wb['分車表 (2)']

# The afternoon columns encode the itinerary each coach follows, which lets us
# disambiguate a name that appears in two coaches of the same department.
AFT_HANSHIN = '(B)洲際漢神踩點巡禮'
AFT_GREEN = '(A)台中綠美圖踏青'
EVE_GAME = '(A)洲際球場觀賽'
EVE_HOME = '(B)賦歸'

def activity_from_header(text):
    t = str(text or '')
    aft = AFT_GREEN if '綠美圖' in t else (AFT_HANSHIN if '漢神' in t else None)
    eve = EVE_HOME if '不看球' in t else (EVE_GAME if '看球' in t else None)
    return (aft, eve)

def parse_bus_table(header_row, unit_row, row_start, row_end, field):
    cols = {}
    for c in range(2, ws3.max_column + 1):
        v = ws3.cell(row=header_row, column=c).value
        if not (v and isinstance(v, str) and '車' in v):
            continue
        label = v.split('\n')[0].strip()
        hint_src = v if unit_row is None else ws3.cell(row=unit_row, column=c).value
        tags = [t for t in ('經策', '數平', '作資') if t in str(hint_src)]
        cols[c] = (label,
                   tags[0] if len(tags) == 1 else None,
                   activity_from_header(v) if unit_row is not None else None)
    matched = 0
    for r in range(row_start, row_end + 1):
        for c, (label, dept_hint, act_hint) in cols.items():
            eid = match_emp_id(ws3.cell(row=r, column=c).value, dept_hint, act_hint,
                               where=f'{field} r{r}c{c}')
            if eid:
                employees[eid][field] = label
                matched += 1
    return matched

print('morning bus matched', parse_bus_table(4, None, 5, 46, 'morning_bus'))
print('activity bus matched', parse_bus_table(48, 49, 50, 90, 'activity_bus'))

# ---------- 5. Table assignment (分桌表) ----------
ws4 = wb['分桌表']
tm = 0
for r in range(4, ws4.max_row + 1):
    table_no = ws4.cell(row=r, column=1).value
    # Skip zone headers ('A區(168)') and the 板前 section marker.
    if not table_no or re.match(r'^[ABC]區', str(table_no)) or '板前' in str(table_no):
        continue
    hint = None
    for tag in ('經策', '數平', '作資'):
        if tag in str(ws4.cell(row=r, column=3).value or ''):
            hint = tag
            break
    for c in range(4, 12):
        eid = match_emp_id(ws4.cell(row=r, column=c).value, hint, where=f'table r{r}c{c}')
        if eid:
            employees[eid]['table'] = str(table_no).strip()
            tm += 1
print('table matched', tm)

# ---------- 6. HSR named seats ----------
hsr_raw = json.load(open(os.path.join(HERE, 'hsr_raw.json')))

def attach_hsr(records, field):
    matched, skipped = 0, 0
    for rec in records:
        # Cells that sit outside any train block (no train number / seat letter)
        # are stray leftovers in the sheet, not real seat assignments. Letting
        # them through wiped out the real seat of anyone who appeared in one.
        if not rec.get('train_no') or not rec.get('seat_letter'):
            skipped += 1
            continue
        hint = None
        for tag in ('經策', '數平', '作資'):
            if tag in str(rec.get('dept_raw') or ''):
                hint = tag
                break
        eid = match_emp_id(rec['name_raw'], hint, where=f'{field} {rec["name_raw"]}')
        if eid:
            employees[eid][field] = {
                'train_no': rec['train_no'], 'car': rec['car'],
                'seat_letter': rec['seat_letter'], 'row_idx': rec['row_idx'],
                'station': rec['station'], 'note': rec['note'],
            }
            matched += 1
    return matched, skipped

print('hsr_outbound matched/skipped', attach_hsr(hsr_raw['go'], 'hsr_outbound'))
print('hsr_return  matched/skipped', attach_hsr(hsr_raw['back'], 'hsr_return'))

# ---------- 7. Write ----------
for e in employees.values():
    e.pop('_tag', None)
out = list(employees.values())
with open(os.path.join(HERE, '..', 'data', 'employees.json'), 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1, default=str)

print(f'\nTOTAL EMPLOYEES {len(out)}')
for f_ in ['morning_bus', 'activity_bus', 'table', 'hsr_outbound', 'hsr_return']:
    print(f'  {f_:<15} {sum(1 for e in out if e[f_])}/{len(out)}')

# Names in the seating sheets that could not be tied to an attending employee.
# These are almost always source-data issues (unnamed '作資' placeholder seats,
# or a person who is in a seating chart but has no registration row), not code
# bugs - but they are printed so 梯長/福委 can check them against the workbook.
real = [u for u in unresolved if u[1].strip() not in ('作資', '經策', '數平', '總務', '空位')]
print(f'\nunresolved names ({len(real)} real, {len(unresolved) - len(real)} unnamed placeholders):')
for u in real:
    print('   ', u)
print(f'\nunmatched dependent rows ({len(dep_unmatched)}) - dependents whose employee '
      'is not in the attending roster (source-data issue):')
for u in dep_unmatched:
    print('   ', u)
